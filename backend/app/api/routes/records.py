from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
import io
from datetime import datetime

from app.schemas.records import ApiMessage, RecordCreateRequest
from app.services.records import create_record, list_records
from app.services.supabase_client import get_supabase


router = APIRouter(tags=["records"])


@router.get("/records", response_model=ApiMessage)
def get_records() -> ApiMessage:
    return ApiMessage(ok=True, message="ok", data={"items": list_records()})


@router.post("/records", response_model=ApiMessage)
def post_record(payload: RecordCreateRequest) -> ApiMessage:
    ok, message, data = create_record(payload.model_dump())
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return ApiMessage(ok=True, message=message, data=data)


@router.get("/records/calendar")
def calendar_data(mes: str, empleado: Optional[str] = None):
    supabase = get_supabase()
    q = supabase.table("registros").select("*").gte("fecha_hora", f"{mes}-01").lte("fecha_hora", f"{mes}-31").execute()
    rows = q.data or []
    days: dict[str, list[str]] = {}
    for r in rows:
        d = r.get("fecha_hora", "")[:10]
        if not d:
            continue
        if empleado and r.get("empleado") != empleado:
            continue
        if d not in days:
            days[d] = []
        days[d].append(r.get("estatus", ""))
    result = {}
    for d, statuses in days.items():
        if any("retardo" in s.lower() for s in statuses):
            result[d] = "retardo"
        elif any(s == "Permiso" for s in statuses):
            result[d] = "permiso"
        elif any(s in ("A Tiempo", "OLVIDO REGISTRO", "Justificado") for s in statuses):
            result[d] = "presente"
        else:
            result[d] = "otro"
    return result


@router.put("/admin/records/{record_id}")
def update_record(record_id: str, payload: dict):
    supabase = get_supabase()
    allowed = {}
    if "estatus" in payload:
        allowed["estatus"] = payload["estatus"]
    if "justificacion" in payload:
        allowed["justificacion"] = payload["justificacion"]
    if not allowed:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = supabase.table("registros").update(allowed).eq("id", record_id).execute()
    if result.data:
        return {"ok": True, "data": result.data[0]}
    raise HTTPException(status_code=404, detail="Record not found")


@router.get("/records/export/excel")
def export_excel(fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None, sucursal_id: Optional[str] = None):
    from openpyxl import Workbook
    supabase = get_supabase()
    q = supabase.table("registros").select("*").order("fecha_hora", desc=True).limit(5000).execute()
    rows = q.data or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    ws.append(["Fecha/Hora", "Empleado", "Tipo", "Estatus", "Min Retardo", "Sucursal", "Justificación"])

    for r in rows:
        fh = r.get("fecha_hora", "")
        if fecha_inicio and fh < fecha_inicio:
            continue
        if fecha_fin and fh > fecha_fin + "T23:59:59":
            continue
        if sucursal_id and str(r.get("sucursal_id", "")) != sucursal_id:
            continue
        ws.append([
            fh,
            r.get("empleado", ""),
            r.get("tipo", ""),
            r.get("estatus", ""),
            r.get("min_retardo", 0),
            str(r.get("sucursal_id", "")),
            r.get("justificacion", ""),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"asistencia_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/records/export/semanal")
def export_semanal(fecha_inicio: str, fecha_fin: str):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    from datetime import date as date_type, timedelta
    import os
    supabase = get_supabase()

    records = (supabase.table("registros").select("*")
        .gte("fecha_hora", f"{fecha_inicio}T00:00:00")
        .lte("fecha_hora", f"{fecha_fin}T23:59:59")
        .execute().data or [])

    empleados = supabase.table("empleados").select("nombre").execute().data or []
    todos_empleados = [e["nombre"] for e in empleados]

    inicio = date_type.fromisoformat(fecha_inicio)
    fin = date_type.fromisoformat(fecha_fin)
    dias = []
    d = inicio
    while d <= fin:
        dias.append(d)
        d += timedelta(days=1)
    dias_semana = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

    emp_diario = defaultdict(lambda: defaultdict(list))
    for r in records:
        try:
            dia = r["fecha_hora"][:10]
        except:
            continue
        emp_diario[r.get("empleado", "")][dia].append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Semanal"

    # Colors
    azul_oscuro = "0a1526"
    azul_header = "1a3a5c"
    cyan = "5ef2ff"
    verde = "9cffb5"
    rojo = "ff8c9e"
    naranja = "ffcc5e"
    gris = "334466"

    title_font = Font(name="Calibri", size=20, bold=True, color=cyan)
    subtitle_font = Font(name="Calibri", size=11, color="aabbcc")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="FFFFFF")
    data_font_green = Font(name="Calibri", size=10, color="003300")
    header_fill = PatternFill(start_color=azul_header, end_color=azul_header, fill_type="solid")
    even_fill = PatternFill(start_color="0d1f33", end_color="0d1f33", fill_type="solid")
    odd_fill = PatternFill(start_color=azul_oscuro, end_color=azul_oscuro, fill_type="solid")
    verde_fill = PatternFill(start_color=verde, end_color=verde, fill_type="solid")
    rojo_fill = PatternFill(start_color=rojo, end_color=rojo, fill_type="solid")
    naranja_fill = PatternFill(start_color=naranja, end_color=naranja, fill_type="solid")
    gris_fill = PatternFill(start_color=gris, end_color=gris, fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="224466"),
        right=Side(style="thin", color="224466"),
        top=Side(style="thin", color="224466"),
        bottom=Side(style="thin", color="224466"),
    )
    no_border = Border()
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Row 1: Logo + Title
    logo_path = os.path.join(os.path.dirname(__file__), "..", "..", "static", "logo.png")
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            img.width = 120
            img.height = 36
            img.anchor = "A1"
            ws.add_image(img)
        except:
            pass

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "  NEOMOTIC"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 42

    # Row 2: Subtitle
    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = f"Reporte Semanal de Asistencia  |  {fecha_inicio}  →  {fecha_fin}"
    sub.font = subtitle_font
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # Row 3: separator
    ws.merge_cells("A3:K3")
    sep = ws["A3"]
    sep.fill = PatternFill(start_color=cyan, end_color=cyan, fill_type="solid")
    ws.row_dimensions[3].height = 3

    # Row 4: empty spacer
    ws.row_dimensions[4].height = 6

    # Row 5: Headers
    header_row = 5
    headers = ["No.", "Empleado", "Días Trab."]
    for di in dias:
        headers.append(dias_semana[di.weekday()])
    headers += ["Retardos", "Min Retardo", "Faltas", "Horas Extra"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 28

    # Data rows
    for idx, emp_name in enumerate(todos_empleados):
        row_num = header_row + 1 + idx
        row_fill = even_fill if idx % 2 == 0 else odd_fill

        trabajados = 0
        total_retardos = 0
        total_ret_min = 0
        faltas = 0
        day_values = []

        for dia in dias:
            key = dia.isoformat()
            day_recs = emp_diario.get(emp_name, {}).get(key, [])
            entrada_recs = [r for r in day_recs if r.get("tipo") == "Entrada"]

            if not entrada_recs:
                if dia.weekday() < 5:
                    day_values.append(("❌", naranja_fill, Font(size=14, color="000000")))
                    faltas += 1
                else:
                    day_values.append(("-", gris_fill, Font(size=14, color="667788")))
                continue

            est = entrada_recs[0].get("estatus", "")
            if "retardo" in est.lower():
                day_values.append(("⚠️", rojo_fill, Font(size=14, color="FFFFFF")))
                trabajados += 1
                total_retardos += 1
                total_ret_min += sum(r.get("min_retardo", 0) for r in entrada_recs if "retardo" in r.get("estatus", "").lower())
            else:
                day_values.append(("✅", verde_fill, Font(size=14, color="003300")))
                trabajados += 1

        row_data = [idx + 1, emp_name, trabajados]
        for val, fill, font in day_values:
            row_data.append(val)
        row_data += [total_retardos, total_ret_min, faltas, ""]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = data_font
            cell.fill = row_fill

        # Override day column styling
        for di, (val, fill, font) in enumerate(day_values):
            ci = 4 + di
            cell = ws.cell(row=row_num, column=ci)
            cell.fill = fill
            cell.font = font
            cell.alignment = center_align
            cell.border = thin_border

        # Employee name left-aligned
        ws.cell(row=row_num, column=2).alignment = left_align

        ws.row_dimensions[row_num].height = 26

    # Column widths
    col_widths = {1: 6, 2: 30, 3: 11}
    for di in range(len(dias)):
        col_widths[4 + di] = 8
    col_widths[4 + len(dias)] = 10     # Retardos
    col_widths[5 + len(dias)] = 12     # Min Retardo
    col_widths[6 + len(dias)] = 8      # Faltas
    col_widths[7 + len(dias)] = 13     # Horas Extra
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Footer
    footer_row = header_row + 1 + len(todos_empleados) + 1
    ws.merge_cells(f"A{footer_row}:K{footer_row}")
    footer = ws.cell(row=footer_row, column=1)
    footer.value = f"Generado por NeoAssistence — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer.font = Font(size=8, color="667788", italic=True)
    footer.alignment = Alignment(horizontal="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"reporte_semanal_{fecha_inicio}_{fecha_fin}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
